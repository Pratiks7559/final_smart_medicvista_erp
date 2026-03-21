from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count, Max, F, Case, When, DecimalField, Value
from django.core.paginator import Paginator
from django.db import connection
from django.http import JsonResponse, HttpResponse
from .models import InventoryTransaction, ProductMaster, Pharmacy_Details, SaleRateMaster
from decimal import Decimal
import time


@login_required
def inventory_list2(request):
    """
    Optimized Inventory list using InventoryTransaction system
    Real-time accurate stock with optimized queries
    """
    
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    stock_filter = request.GET.get('stock_filter', 'all')
    page_number = request.GET.get('page', 1)
    
    # OPTIMIZED: Single aggregated query with prefetch
    stock_summary = InventoryTransaction.objects.select_related(
        'product'
    ).values(
        'product_id',
        'product__product_name',
        'product__product_company',
        'product__product_packing',
        'product__product_salt',
        'product__product_category'
    ).annotate(
        total_stock=Sum('quantity'),
        total_free=Sum('free_quantity'),
        batch_count=Count('batch_no', distinct=True),
        avg_mrp=Max('mrp')
    )  # Remove filter to show all products with batches, even if stock is 0
    
    # Apply search filter
    if search_query:
        stock_summary = stock_summary.filter(
            Q(product__product_name__icontains=search_query) |
            Q(product__product_company__icontains=search_query) |
            Q(product__product_salt__icontains=search_query) |
            Q(product__product_category__icontains=search_query)
        )
    
    # Convert to list for processing
    stock_data = list(stock_summary)
    
    # OPTIMIZATION: Bulk fetch all product IDs
    product_ids = [item['product_id'] for item in stock_data]
    
    # OPTIMIZATION: Prefetch all batches in one query (including zero stock)
    all_batches = InventoryTransaction.objects.filter(
        product_id__in=product_ids
    ).values(
        'product_id', 'batch_no', 'expiry_date', 'mrp'
    ).annotate(
        stock=Sum('quantity'),
        free_stock=Sum('free_quantity')
    ).order_by('product_id', 'expiry_date', 'batch_no')  # Show all batches including zero stock
    
    # OPTIMIZATION: Prefetch all rates in one query
    all_rates = SaleRateMaster.objects.filter(
        productid_id__in=product_ids
    ).select_related('productid').values(
        'productid_id', 'product_batch_no', 'rate_A', 'rate_B', 'rate_C'
    )
    
    # Create lookup dictionaries for O(1) access
    batches_by_product = {}
    for batch in all_batches:
        pid = batch['product_id']
        if pid not in batches_by_product:
            batches_by_product[pid] = []
        batches_by_product[pid].append(batch)
    
    rates_lookup = {}
    for rate in all_rates:
        key = f"{rate['productid_id']}_{rate['product_batch_no']}"
        rates_lookup[key] = {
            'rate_A': rate['rate_A'],
            'rate_B': rate['rate_B'],
            'rate_C': rate['rate_C']
        }
    
    # Process inventory data
    inventory_data = []
    summary_stats = {'in_stock': 0, 'low_stock': 0, 'out_of_stock': 0, 'total_value': 0}
    
    for item in stock_data:
        product_id = item['product_id']
        batches = batches_by_product.get(product_id, [])
        
        batches_info = []
        total_stock = 0
        stock_value = 0
        
        for batch in batches:
            batch_stock = (batch['stock'] or 0) + (batch['free_stock'] or 0)
            total_stock += batch_stock
            stock_value += (batch['stock'] or 0) * (batch['mrp'] or 0)
            
            # Get rates from lookup
            rate_key = f"{product_id}_{batch['batch_no']}"
            rates = rates_lookup.get(rate_key, {'rate_A': 0, 'rate_B': 0, 'rate_C': 0})
            
            batches_info.append({
                'batch_no': batch['batch_no'],
                'expiry': batch['expiry_date'],
                'stock': batch_stock,
                'qty': batch['stock'] or 0,
                'free_qty': batch['free_stock'] or 0,
                'mrp': batch['mrp'] or 0,
                'rates': rates
            })
        
        # Determine status
        if total_stock <= 0:
            status = 'Out of Stock'
            summary_stats['out_of_stock'] += 1
        elif total_stock < 10:
            status = 'Low Stock'
            summary_stats['low_stock'] += 1
        else:
            status = 'In Stock'
            summary_stats['in_stock'] += 1
        
        summary_stats['total_value'] += stock_value
        
        # Apply stock filter
        if stock_filter == 'in_stock' and total_stock < 10:
            continue
        elif stock_filter == 'low_stock' and (total_stock <= 0 or total_stock >= 10):
            continue
        elif stock_filter == 'out_of_stock' and total_stock > 0:
            continue
        
        inventory_data.append({
            'product': ProductMaster.objects.get(productid=product_id),
            'total_stock': total_stock,
            'stock_value': stock_value,
            'status': status,
            'batches_info': batches_info
        })
    
    # Manual pagination
    offset = int(request.GET.get('offset', 0))
    per_page = 50
    start = offset
    end = offset + per_page
    
    inventory_slice = inventory_data[start:end]
    total_count = len(inventory_data)
    has_more = end < total_count
    next_offset = end if has_more else None
    
    # Page statistics
    page_total_value = sum(item['stock_value'] for item in inventory_slice)
    page_low_stock = sum(1 for item in inventory_slice if item['status'] == 'Low Stock')
    page_out_of_stock = sum(1 for item in inventory_slice if item['status'] == 'Out of Stock')
    
    # Check if AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        from django.template.loader import render_to_string
        
        html = render_to_string('inventory/inventory_rows.html', {'inventory_data': inventory_slice})
        
        return JsonResponse({
            'success': True,
            'html': html,
            'has_more': has_more,
            'next_offset': next_offset,
            'total_count': total_count,
            'batch_stats': {
                'total_value': page_total_value,
                'low_stock': page_low_stock,
                'out_of_stock': page_out_of_stock
            }
        })
    
    context = {
        'inventory_data': inventory_slice,
        'search_query': search_query,
        'total_products': total_count,
        'page_total_value': page_total_value,
        'page_low_stock': page_low_stock,
        'page_out_of_stock': page_out_of_stock,
        'has_more': has_more,
        'next_offset': next_offset,
        'pharmacy': Pharmacy_Details.objects.first(),
        'title': 'Inventory List 2 (Real-time)',
    }
    
    return render(request, 'inventory/inventory_list2.html', context)


@login_required
def inventory_batches(request, product_id):
    """
    Get batch details for a product (AJAX endpoint)
    """
    from django.http import JsonResponse
    
    batches = InventoryTransaction.objects.filter(
        product_id=product_id
    ).values(
        'batch_no', 'expiry_date', 'mrp'
    ).annotate(
        total_qty=Sum('quantity'),
        total_free_qty=Sum('free_quantity'),
        last_rate=Max('rate')
    ).filter(
        Q(total_qty__gt=0) | Q(total_free_qty__gt=0)
    ).order_by('expiry_date', 'batch_no')
    
    batch_list = []
    for batch in batches:
        batch_list.append({
            'batch_no': batch['batch_no'],
            'expiry_date': batch['expiry_date'],
            'mrp': float(batch['mrp']),
            'stock': float(batch['total_qty'] or 0),
            'free_stock': float(batch['total_free_qty'] or 0),
            'total_stock': float((batch['total_qty'] or 0) + (batch['total_free_qty'] or 0)),
            'last_rate': float(batch['last_rate'] or 0)
        })
    
    return JsonResponse({'batches': batch_list})


@login_required
def inventory_transaction_history(request, product_id):
    """View transaction history for a product with pagination"""
    
    product = get_object_or_404(ProductMaster, productid=product_id)
    batch_no = request.GET.get('batch_no', '')
    
    # Get transactions
    transactions = InventoryTransaction.objects.filter(product=product)
    
    if batch_no:
        transactions = transactions.filter(batch_no=batch_no)
    
    transactions = transactions.select_related('created_by').order_by('-transaction_date')
    
    # Pagination - 15 rows per page
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(transactions, 15)  # 15 transactions per page
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    # Calculate running balance for current page
    transaction_list = []
    
    if page_obj:
        # Get all transactions in chronological order for balance calculation
        all_txns = list(InventoryTransaction.objects.filter(
            product=product
        ).order_by('transaction_date', 'transaction_id'))
        
        if batch_no:
            all_txns = [t for t in all_txns if t.batch_no == batch_no]
        
        # Create balance lookup
        balance_lookup = {}
        running_balance = Decimal('0')
        
        for txn in all_txns:
            running_balance += Decimal(str(txn.quantity))
            balance_lookup[txn.transaction_id] = running_balance
        
        # Add balance to current page transactions
        for txn in page_obj:
            transaction_list.append({
                'transaction': txn,
                'balance': balance_lookup.get(txn.transaction_id, 0)
            })
    
    context = {
        'product': product,
        'batch_no': batch_no,
        'transactions': transaction_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'title': f'Transaction History - {product.product_name}'
    }
    
    return render(request, 'inventory/transaction_history.html', context)


@login_required
def inventory_dashboard(request):
    """Dashboard with inventory statistics"""
    
    # Get summary statistics using optimized queries
    stats = InventoryTransaction.objects.aggregate(
        total_stock=Sum('quantity'),
        total_free=Sum('free_quantity'),
        total_transactions=Count('transaction_id')
    )
    
    # Get products with stock
    products_with_stock = InventoryTransaction.objects.values('product_id').annotate(
        total=Sum('quantity') + Sum('free_quantity')
    ).filter(total__gt=0).count()
    
    # Get low stock products
    low_stock_products = InventoryTransaction.objects.values('product_id').annotate(
        total=Sum('quantity') + Sum('free_quantity')
    ).filter(total__lte=10, total__gt=0).count()
    
    # Get out of stock products
    all_products = ProductMaster.objects.count()
    out_of_stock = all_products - products_with_stock
    
    context = {
        'total_stock': stats['total_stock'] or 0,
        'total_free': stats['total_free'] or 0,
        'total_transactions': stats['total_transactions'],
        'products_with_stock': products_with_stock,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock,
        'title': 'Inventory Dashboard'
    }
    
    return render(request, 'inventory/dashboard.html', context)



@login_required
def export_inventory2_pdf(request):
    """Export inventory list 2 to PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import io
    from datetime import datetime
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Get search query
        search_query = request.GET.get('search', '').strip()
        stock_filter = request.GET.get('stock_filter', 'all')
        
        # Get inventory data using same logic as main view
        stock_summary = InventoryTransaction.objects.values(
            'product_id',
            'product__product_name',
            'product__product_company',
            'product__product_packing',
            'product__product_category'
        ).annotate(
            total_stock=Sum('quantity'),
            total_free=Sum('free_quantity'),
            batch_count=Count('batch_no', distinct=True),
            avg_mrp=Max('mrp')
        ).filter(Q(total_stock__gt=0) | Q(total_free__gt=0))
        
        if search_query:
            stock_summary = stock_summary.filter(
                Q(product__product_name__icontains=search_query) |
                Q(product__product_company__icontains=search_query) |
                Q(product__product_salt__icontains=search_query)
            )
        
        stock_data = list(stock_summary)
        
        # Process inventory data
        inventory_data = []
        total_value = 0
        
        for item in stock_data:
            batches = InventoryTransaction.objects.filter(
                product_id=item['product_id']
            ).values('batch_no', 'expiry_date', 'mrp').annotate(
                stock=Sum('quantity'),
                free_stock=Sum('free_quantity')
            ).filter(Q(stock__gt=0) | Q(free_stock__gt=0))
            
            total_stock = 0
            stock_value = 0
            
            for batch in batches:
                batch_stock = (batch['stock'] or 0) + (batch['free_stock'] or 0)
                total_stock += batch_stock
                stock_value += (batch['stock'] or 0) * (batch['mrp'] or 0)
            
            # Apply stock filter
            if stock_filter == 'in_stock' and total_stock < 10:
                continue
            elif stock_filter == 'low_stock' and (total_stock <= 0 or total_stock >= 10):
                continue
            elif stock_filter == 'out_of_stock' and total_stock > 0:
                continue
            
            inventory_data.append({
                'product_name': item['product__product_name'],
                'company': item['product__product_company'],
                'packing': item['product__product_packing'],
                'category': item['product__product_category'] or 'N/A',
                'total_stock': total_stock,
                'stock_value': stock_value,
                'batch_count': item['batch_count']
            })
            total_value += stock_value
        
        # Pharmacy header
        try:
            pharmacy = Pharmacy_Details.objects.first()
            if pharmacy:
                title_style = styles['Heading1']
                title_style.alignment = 1
                if pharmacy.pharmaname:
                    story.append(Paragraph(pharmacy.pharmaname.upper(), title_style))
                info_style = styles['Normal']
                info_style.alignment = 1
                info_style.fontSize = 9
                pharmacy_info = []
                if pharmacy.proprietorcontact:
                    pharmacy_info.append(f"Contact: {pharmacy.proprietorcontact}")
                if pharmacy.proprietoremail:
                    pharmacy_info.append(f"Email: {pharmacy.proprietoremail}")
                if pharmacy_info:
                    story.append(Paragraph(" | ".join(pharmacy_info), info_style))
                story.append(Spacer(1, 0.15*inch))
        except:
            pass
        
        # Title
        title_style = styles['Heading1']
        title_style.alignment = 1
        title = Paragraph("INVENTORY REPORT (Real-time)", title_style)
        story.append(title)
        
        # Date
        date_style = styles['Normal']
        date_style.alignment = 1
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}", date_style)
        story.append(date_text)
        story.append(Spacer(1, 0.3*inch))
        
        # Summary
        total_products = len(inventory_data)
        out_of_stock = sum(1 for item in inventory_data if item['total_stock'] <= 0)
        low_stock = sum(1 for item in inventory_data if 0 < item['total_stock'] < 10)
        
        summary_data = [
            ['Total Products', 'Total Value', 'Out of Stock', 'Low Stock'],
            [str(total_products), f"₹{total_value:,.2f}", str(out_of_stock), str(low_stock)]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2.5*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Inventory table
        table_data = [['Product Name', 'Company', 'Packing', 'Category', 'Batches', 'Stock', 'Value']]
        
        for item in inventory_data:
            table_data.append([
                item['product_name'][:25] + '...' if len(item['product_name']) > 25 else item['product_name'],
                item['company'][:15] + '...' if len(item['company']) > 15 else item['company'],
                item['packing'][:10] + '...' if len(item['packing']) > 10 else item['packing'],
                item['category'][:10] + '...' if len(item['category']) > 10 else item['category'],
                str(item['batch_count']),
                str(int(item['total_stock'])),
                f"₹{item['stock_value']:.0f}"
            ])
        
        inventory_table = Table(table_data, repeatRows=1)
        inventory_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONT', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (6, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(inventory_table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory_report_realtime.pdf"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
def export_inventory2_excel(request):
    """Export inventory list 2 to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    import io
    
    try:
        # Get search query
        search_query = request.GET.get('search', '').strip()
        stock_filter = request.GET.get('stock_filter', 'all')
        
        # Get inventory data
        stock_summary = InventoryTransaction.objects.values(
            'product_id',
            'product__product_name',
            'product__product_company',
            'product__product_packing',
            'product__product_category'
        ).annotate(
            total_stock=Sum('quantity'),
            total_free=Sum('free_quantity'),
            batch_count=Count('batch_no', distinct=True),
            avg_mrp=Max('mrp')
        ).filter(Q(total_stock__gt=0) | Q(total_free__gt=0))
        
        if search_query:
            stock_summary = stock_summary.filter(
                Q(product__product_name__icontains=search_query) |
                Q(product__product_company__icontains=search_query) |
                Q(product__product_salt__icontains=search_query)
            )
        
        stock_data = list(stock_summary)
        
        # Process data
        inventory_data = []
        total_value = 0
        
        for item in stock_data:
            batches = InventoryTransaction.objects.filter(
                product_id=item['product_id']
            ).values('batch_no', 'expiry_date', 'mrp').annotate(
                stock=Sum('quantity'),
                free_stock=Sum('free_quantity')
            ).filter(Q(stock__gt=0) | Q(free_stock__gt=0))
            
            total_stock = 0
            stock_value = 0
            
            for batch in batches:
                batch_stock = (batch['stock'] or 0) + (batch['free_stock'] or 0)
                total_stock += batch_stock
                stock_value += (batch['stock'] or 0) * (batch['mrp'] or 0)
            
            # Apply stock filter
            if stock_filter == 'in_stock' and total_stock < 10:
                continue
            elif stock_filter == 'low_stock' and (total_stock <= 0 or total_stock >= 10):
                continue
            elif stock_filter == 'out_of_stock' and total_stock > 0:
                continue
            
            inventory_data.append({
                'product_name': item['product__product_name'],
                'company': item['product__product_company'],
                'packing': item['product__product_packing'],
                'category': item['product__product_category'] or 'N/A',
                'total_stock': total_stock,
                'stock_value': stock_value,
                'batch_count': item['batch_count']
            })
            total_value += stock_value
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Pharmacy header
        try:
            pharmacy = Pharmacy_Details.objects.first()
            if pharmacy:
                ws['A1'] = pharmacy.pharmaname or "Pharmacy"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A1:G1')
                
                ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}"
                ws['A2'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A2:G2')
                
                current_row = 4
            else:
                current_row = 1
        except:
            current_row = 1
        
        # Title
        ws[f'A{current_row}'] = "INVENTORY REPORT (Real-time)"
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2
        
        # Summary
        ws[f'A{current_row}'] = "Total Products"
        ws[f'B{current_row}'] = len(inventory_data)
        ws[f'C{current_row}'] = "Total Value"
        ws[f'D{current_row}'] = f"₹{total_value:,.2f}"
        current_row += 2
        
        # Headers
        headers = ['Product Name', 'Company', 'Packing', 'Category', 'Batches', 'Stock', 'Value']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        current_row += 1
        
        # Data rows
        for item in inventory_data:
            ws.cell(row=current_row, column=1, value=item['product_name']).border = border
            ws.cell(row=current_row, column=2, value=item['company']).border = border
            ws.cell(row=current_row, column=3, value=item['packing']).border = border
            ws.cell(row=current_row, column=4, value=item['category']).border = border
            ws.cell(row=current_row, column=5, value=item['batch_count']).border = border
            ws.cell(row=current_row, column=6, value=int(item['total_stock'])).border = border
            ws.cell(row=current_row, column=7, value=f"₹{item['stock_value']:.2f}").border = border
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_report_realtime.xlsx"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)

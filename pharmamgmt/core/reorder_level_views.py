from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from collections import namedtuple
import pandas as pd
from .models import ProductMaster, SalesMaster, BatchInventoryCache, Pharmacy_Details

ReorderStats = namedtuple('ReorderStats', ['avg_monthly_sale', 'reorder_level', 'total_available', 'reorder_needed'])

# ── Tunable constants ────────────────────────────────────────────────────────
ANALYSIS_DAYS  = 90   # sales look-back window (3 months → stable average)
LEAD_TIME_DAYS = 30   # reorder level = stock needed to cover 1 month lead time
# ─────────────────────────────────────────────────────────────────────────────


def _product_reorder_stats(product, batches, analysis_start):
    """
    Correct reorder calculations:

    avg_daily_sale   = total_sales_in_window / ANALYSIS_DAYS
    avg_monthly_sale = avg_daily_sale * 30          (display only)
    reorder_level    = avg_daily_sale * LEAD_TIME_DAYS
                       → minimum stock that must be on hand before placing order
    total_available  = sum of current_stock across all batches (free qty excluded)
    reorder_needed   = max(0, reorder_level - total_available)
                       → how many units to order right now
    """
    sales_qty = SalesMaster.objects.filter(
        productid=product,
        sale_entry_date__gte=analysis_start
    ).aggregate(total=Sum('sale_quantity'))['total'] or 0

    total_sales      = float(sales_qty)
    avg_daily_sale   = total_sales / ANALYSIS_DAYS
    avg_monthly_sale = round(avg_daily_sale * 30, 2)
    reorder_level    = round(avg_daily_sale * LEAD_TIME_DAYS, 2)
    total_available  = round(sum(float(b.current_stock) for b in batches), 2)
    reorder_needed   = round(max(0.0, reorder_level - total_available), 2)

    return ReorderStats(avg_monthly_sale, reorder_level, total_available, reorder_needed)


@login_required
def reorder_level_report(request):
    """Display product-wise and batch-wise reorder levels with pagination"""

    product_search    = request.GET.get('product_search', '')
    show_reorder_only = request.GET.get('show_reorder_only') == 'true'
    page_number       = request.GET.get('page', 1)

    products_query = ProductMaster.objects.prefetch_related('batch_caches').all()

    if product_search:
        products_query = products_query.filter(
            Q(product_name__icontains=product_search) |
            Q(product_company__icontains=product_search)
        )

    analysis_start = datetime.now() - timedelta(days=ANALYSIS_DAYS)
    reorder_data   = []

    for product in products_query:
        batches = product.batch_caches.filter(current_stock__gt=0).order_by('expiry_date', 'batch_no')
        if not batches.exists():
            continue

        avg_monthly_sale, reorder_level, total_available, reorder_needed = \
            _product_reorder_stats(product, batches, analysis_start)

        if show_reorder_only and reorder_needed <= 0:
            continue

        batch_details = []
        for batch in batches:
            batch_details.append({
                'batch_no':       batch.batch_no,
                'expiry_date':    batch.expiry_date,
                'mrp':            batch.mrp,
                'purchase_rate':  batch.purchase_rate,
                'available_stock': float(batch.current_stock),
                'free_qty':       float(batch.current_free_qty),
                'rate_a':         batch.rate_a,
                'rate_b':         batch.rate_b,
                'rate_c':         batch.rate_c,
            })

        reorder_data.append({
            'product_id':      product.productid,
            'product_name':    product.product_name,
            'product_company': product.product_company,
            'product_packing': product.product_packing,
            'avg_monthly_sale': avg_monthly_sale,
            'reorder_level':   reorder_level,
            'total_available': total_available,
            'reorder_needed':  reorder_needed,
            'batches':         batch_details,
            'status':          'critical' if reorder_needed > 0 else 'sufficient',
        })

    reorder_data.sort(key=lambda x: x['reorder_needed'], reverse=True)

    paginator = Paginator(reorder_data, 20)
    page_obj  = paginator.get_page(page_number)

    context = {
        'title':            'Reorder Level Report',
        'page_obj':         page_obj,
        'product_search':   product_search,
        'show_reorder_only': show_reorder_only,
        'pharmacy':         Pharmacy_Details.objects.first(),
        'total_products':   len(reorder_data),
        'analysis_days':    ANALYSIS_DAYS,
        'lead_time_days':   LEAD_TIME_DAYS,
    }
    return render(request, 'purchases/reorder_level_report.html', context)


@login_required
def export_reorder_level_excel(request):
    """Export reorder level report as Excel"""

    product_search    = request.GET.get('product_search', '')
    show_reorder_only = request.GET.get('show_reorder_only') == 'true'

    products_query = ProductMaster.objects.all()
    if product_search:
        products_query = products_query.filter(
            Q(product_name__icontains=product_search) |
            Q(product_company__icontains=product_search)
        )

    analysis_start = datetime.now() - timedelta(days=ANALYSIS_DAYS)
    data = []

    for product in products_query:
        batches = BatchInventoryCache.objects.filter(
            product=product, current_stock__gt=0
        ).order_by('expiry_date', 'batch_no')

        if not batches.exists():
            continue

        avg_monthly_sale, reorder_level, total_available, reorder_needed = \
            _product_reorder_stats(product, batches, analysis_start)

        if show_reorder_only and reorder_needed <= 0:
            continue

        for batch in batches:
            data.append({
                'Product Name':              product.product_name,
                'Company':                   product.product_company,
                'Packing':                   product.product_packing,
                'Batch No':                  batch.batch_no,
                'Expiry':                    batch.expiry_date,
                'MRP':                       batch.mrp,
                'Purchase Rate':             batch.purchase_rate,
                'Batch Stock':               float(batch.current_stock),
                'Free Qty':                  float(batch.current_free_qty),
                'Total Available':           total_available,
                f'Avg Monthly Sale ({ANALYSIS_DAYS}d)': avg_monthly_sale,
                f'Reorder Level ({LEAD_TIME_DAYS}d lead)': reorder_level,
                'Reorder Needed':            reorder_needed,
                'Rate A':                    batch.rate_a,
                'Rate B':                    batch.rate_b,
                'Rate C':                    batch.rate_c,
            })

    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reorder_level_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Reorder Level', index=False, startrow=3)

        workbook  = writer.book
        worksheet = writer.sheets['Reorder Level']

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        pharmacy = Pharmacy_Details.objects.first()
        if pharmacy:
            worksheet['A1'] = pharmacy.pharmaname.upper()
            worksheet['A1'].font      = Font(size=16, bold=True)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:P1')

        worksheet['A2'] = (
            f'REORDER LEVEL REPORT - {datetime.now().strftime("%d-%m-%Y")} '
            f'(Sales window: {ANALYSIS_DAYS} days | Lead time: {LEAD_TIME_DAYS} days)'
        )
        worksheet['A2'].font      = Font(size=12, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A2:P2')

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        for col in range(1, 17):
            cell = worksheet.cell(row=4, column=col)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        column_widths = [25, 20, 10, 15, 10, 10, 12, 12, 10, 12, 18, 20, 15, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

    return response

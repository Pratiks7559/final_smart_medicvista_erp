from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Q
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    SalesMaster, PurchaseMaster, SupplierChallanMaster,
    CustomerChallanMaster, ReturnSalesMaster, ReturnPurchaseMaster
)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: build filtered querysets using ACTUAL invoice/challan dates
# ─────────────────────────────────────────────────────────────────────────────
def _build_queries(start_date, end_date, product_id, product_search):
    # Normalize - strip whitespace
    start_date = (start_date or '').strip()
    end_date   = (end_date   or '').strip()
    sales_q = SalesMaster.objects.select_related(
        'productid', 'sales_invoice_no', 'customerid'
    ).order_by('-sales_invoice_no__sales_invoice_date')

    purchase_q = PurchaseMaster.objects.select_related(
        'productid', 'product_invoiceid', 'product_supplierid'
    ).order_by('-product_invoiceid__invoice_date')

    sup_challan_q = SupplierChallanMaster.objects.select_related(
        'product_id', 'product_suppliername', 'product_challan_id'
    ).filter(product_challan_id__is_invoiced=False
    ).order_by('-product_challan_id__challan_date')

    cust_challan_q = CustomerChallanMaster.objects.select_related(
        'product_id', 'customer_name', 'customer_challan_id'
    ).order_by('-customer_challan_id__customer_challan_date')

    sales_ret_q = ReturnSalesMaster.objects.select_related(
        'return_productid', 'return_sales_invoice_no', 'return_customerid'
    ).order_by('-return_sales_invoice_no__return_sales_invoice_date')

    pur_ret_q = ReturnPurchaseMaster.objects.select_related(
        'returnproductid', 'returninvoiceid', 'returnproduct_supplierid'
    ).order_by('-returninvoiceid__returninvoice_date')

    # ── Date filters using ACTUAL document dates ──────────────────────────
    if start_date:
        sales_q        = sales_q.filter(sales_invoice_no__sales_invoice_date__gte=start_date)
        purchase_q     = purchase_q.filter(product_invoiceid__invoice_date__gte=start_date)
        sup_challan_q  = sup_challan_q.filter(product_challan_id__challan_date__gte=start_date)
        cust_challan_q = cust_challan_q.filter(customer_challan_id__customer_challan_date__gte=start_date)
        sales_ret_q    = sales_ret_q.filter(return_sales_invoice_no__return_sales_invoice_date__gte=start_date)
        pur_ret_q      = pur_ret_q.filter(returninvoiceid__returninvoice_date__gte=start_date)

    if end_date:
        sales_q        = sales_q.filter(sales_invoice_no__sales_invoice_date__lte=end_date)
        purchase_q     = purchase_q.filter(product_invoiceid__invoice_date__lte=end_date)
        sup_challan_q  = sup_challan_q.filter(product_challan_id__challan_date__lte=end_date)
        cust_challan_q = cust_challan_q.filter(customer_challan_id__customer_challan_date__lte=end_date)
        sales_ret_q    = sales_ret_q.filter(return_sales_invoice_no__return_sales_invoice_date__lte=end_date)
        pur_ret_q      = pur_ret_q.filter(returninvoiceid__returninvoice_date__lte=end_date)

    # ── Product filter ────────────────────────────────────────────────────
    if product_id and str(product_id).strip():
        try:
            pid = int(product_id)
            sales_q        = sales_q.filter(productid_id=pid)
            purchase_q     = purchase_q.filter(productid_id=pid)
            sup_challan_q  = sup_challan_q.filter(product_id=pid)
            cust_challan_q = cust_challan_q.filter(product_id=pid)
            sales_ret_q    = sales_ret_q.filter(return_productid_id=pid)
            pur_ret_q      = pur_ret_q.filter(returnproductid_id=pid)
        except (ValueError, TypeError):
            pass
    elif product_search and product_search.strip():
        t = product_search.strip()
        sales_q        = sales_q.filter(product_name__icontains=t)
        purchase_q     = purchase_q.filter(product_name__icontains=t)
        sup_challan_q  = sup_challan_q.filter(product_name__icontains=t)
        cust_challan_q = cust_challan_q.filter(product_name__icontains=t)
        sales_ret_q    = sales_ret_q.filter(return_product_name__icontains=t)
        pur_ret_q      = pur_ret_q.filter(returnproductid__product_name__icontains=t)

    # ── Limit if no date range ────────────────────────────────────────────
    if not (start_date and end_date):
        sales_q        = sales_q[:500]
        purchase_q     = purchase_q[:500]
        sup_challan_q  = sup_challan_q[:500]
        cust_challan_q = cust_challan_q[:500]
        sales_ret_q    = sales_ret_q[:500]
        pur_ret_q      = pur_ret_q[:500]

    return (
        list(sales_q), list(purchase_q),
        list(sup_challan_q), list(cust_challan_q),
        list(sales_ret_q), list(pur_ret_q)
    )


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: build purchase-rate lookup for sale/return rows
# ─────────────────────────────────────────────────────────────────────────────
def _build_purchase_lookup(sales_list, cust_challan_list, sales_ret_list):
    batch_keys = set()
    for s in sales_list:
        batch_keys.add((s.productid_id, s.product_batch_no))
    for c in cust_challan_list:
        batch_keys.add((c.product_id_id, c.product_batch_no))
    for r in sales_ret_list:
        batch_keys.add((r.return_productid_id, r.return_product_batch_no))

    lookup = {}
    if batch_keys:
        q = Q()
        for pid, bn in batch_keys:
            q |= Q(productid_id=pid, product_batch_no=bn)
        for p in PurchaseMaster.objects.filter(q).values('productid_id', 'product_batch_no', 'product_purchase_rate'):
            key = (p['productid_id'], p['product_batch_no'])
            if key not in lookup:
                lookup[key] = float(p['product_purchase_rate'])
    return lookup


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: convert all 6 lists → flat financial_data list + summary totals
# ─────────────────────────────────────────────────────────────────────────────
def _build_financial_data(sales_list, purchase_list, sup_challan_list,
                           cust_challan_list, sales_ret_list, pur_ret_list):
    lookup = _build_purchase_lookup(sales_list, cust_challan_list, sales_ret_list)
    rows = []
    total_sales = total_purchase = total_gst = total_profit = 0.0

    # Sales
    for s in sales_list:
        try:
            pr    = lookup.get((s.productid_id, s.product_batch_no), 0.0)
            qty   = float(s.sale_quantity)
            sr    = float(s.sale_rate)
            cgst  = float(s.sale_cgst); sgst = float(s.sale_sgst)
            sv    = sr * qty
            pc    = pr * qty
            gst   = sv * (cgst + sgst) / 100
            pft   = sv - pc
            rows.append({'type':'Sale',
                'date': s.sales_invoice_no.sales_invoice_date,
                'invoice_no': s.sales_invoice_no.sales_invoice_no,
                'customer': s.customerid.customer_name,
                'product_name': s.product_name, 'company': s.product_company,
                'batch_no': s.product_batch_no, 'quantity': qty,
                'mrp': float(s.product_MRP), 'purchase_rate': pr,
                'sale_rate': sr, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': gst, 'purchase_cost': pc,
                'sales_value': sv, 'profit': pft,
                'profit_percentage': (pft/sv*100) if sv else 0})
            total_sales += sv; total_purchase += pc
            total_gst += gst; total_profit += pft
        except Exception:
            continue

    # Customer Challans
    for c in cust_challan_list:
        try:
            pr    = lookup.get((c.product_id_id, c.product_batch_no), 0.0)
            qty   = float(c.sale_quantity)
            sr    = float(c.sale_rate)
            cgst  = float(c.sale_cgst); sgst = float(c.sale_sgst)
            sv    = sr * qty
            pc    = pr * qty
            gst   = sv * (cgst + sgst) / 100
            pft   = sv - pc
            rows.append({'type':'Customer Challan',
                'date': c.customer_challan_id.customer_challan_date,
                'invoice_no': c.customer_challan_no,
                'customer': c.customer_name.customer_name,
                'product_name': c.product_name, 'company': c.product_company,
                'batch_no': c.product_batch_no, 'quantity': qty,
                'mrp': float(c.product_mrp), 'purchase_rate': pr,
                'sale_rate': sr, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': gst, 'purchase_cost': pc,
                'sales_value': sv, 'profit': pft,
                'profit_percentage': (pft/sv*100) if sv else 0})
            total_sales += sv; total_purchase += pc
            total_gst += gst; total_profit += pft
        except Exception:
            continue

    # Purchases
    for p in purchase_list:
        try:
            qty  = float(p.product_quantity)
            pr   = float(p.product_purchase_rate)
            cgst = float(p.CGST); sgst = float(p.SGST)
            pc   = pr * qty
            gst  = pc * (cgst + sgst) / 100
            rows.append({'type':'Purchase',
                'date': p.product_invoiceid.invoice_date,
                'invoice_no': p.product_invoice_no,
                'customer': p.product_supplierid.supplier_name,
                'product_name': p.product_name, 'company': p.product_company,
                'batch_no': p.product_batch_no, 'quantity': qty,
                'mrp': float(p.product_MRP), 'purchase_rate': pr,
                'sale_rate': 0.0, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': gst, 'purchase_cost': pc,
                'sales_value': 0.0, 'profit': -pc,
                'profit_percentage': 0.0})
            total_purchase += pc; total_gst += gst; total_profit -= pc
        except Exception:
            continue

    # Supplier Challans  ← use challan_date (actual document date)
    for c in sup_challan_list:
        try:
            qty  = float(c.product_quantity)
            pr   = float(c.product_purchase_rate)
            cgst = float(c.cgst); sgst = float(c.sgst)
            pc   = pr * qty
            gst  = pc * (cgst + sgst) / 100
            rows.append({'type':'Supplier Challan',
                'date': c.product_challan_id.challan_date,
                'invoice_no': c.product_challan_no,
                'customer': c.product_suppliername.supplier_name,
                'product_name': c.product_name, 'company': c.product_company,
                'batch_no': c.product_batch_no, 'quantity': qty,
                'mrp': float(c.product_mrp), 'purchase_rate': pr,
                'sale_rate': 0.0, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': gst, 'purchase_cost': pc,
                'sales_value': 0.0, 'profit': -pc,
                'profit_percentage': 0.0})
            total_purchase += pc; total_gst += gst; total_profit -= pc
        except Exception:
            continue

    # Sales Returns
    for r in sales_ret_list:
        try:
            pr   = lookup.get((r.return_productid_id, r.return_product_batch_no), 0.0)
            qty  = float(r.return_sale_quantity)
            sr   = float(r.return_sale_rate)
            cgst = float(r.return_sale_cgst); sgst = float(r.return_sale_sgst)
            sv   = sr * qty
            pc   = pr * qty
            gst  = sv * (cgst + sgst) / 100
            pft  = -(sv - pc)
            rows.append({'type':'Sales Return',
                'date': r.return_sales_invoice_no.return_sales_invoice_date,
                'invoice_no': r.return_sales_invoice_no.return_sales_invoice_no,
                'customer': r.return_customerid.customer_name,
                'product_name': r.return_product_name, 'company': r.return_product_company,
                'batch_no': r.return_product_batch_no, 'quantity': -qty,
                'mrp': float(r.return_product_MRP), 'purchase_rate': pr,
                'sale_rate': sr, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': -gst, 'purchase_cost': -pc,
                'sales_value': -sv, 'profit': pft,
                'profit_percentage': (pft/sv*100) if sv else 0})
            total_sales -= sv; total_purchase -= pc
            total_gst -= gst; total_profit += pft
        except Exception:
            continue

    # Purchase Returns
    for r in pur_ret_list:
        try:
            qty  = float(r.returnproduct_quantity)
            pr   = float(r.returnproduct_purchase_rate)
            cgst = float(r.returnproduct_cgst); sgst = float(r.returnproduct_sgst)
            pc   = pr * qty
            gst  = pc * (cgst + sgst) / 100
            rows.append({'type':'Purchase Return',
                'date': r.returninvoiceid.returninvoice_date,
                'invoice_no': r.returninvoiceid.returninvoiceid,
                'customer': r.returnproduct_supplierid.supplier_name,
                'product_name': r.returnproductid.product_name,
                'company': r.returnproductid.product_company,
                'batch_no': r.returnproduct_batch_no, 'quantity': -qty,
                'mrp': float(r.returnproduct_MRP), 'purchase_rate': pr,
                'sale_rate': 0.0, 'cgst': cgst, 'sgst': sgst,
                'gst_amount': -gst, 'purchase_cost': -pc,
                'sales_value': 0.0, 'profit': pc,
                'profit_percentage': 0.0})
            total_purchase -= pc; total_gst -= gst; total_profit += pc
        except Exception:
            continue

    rows.sort(key=lambda x: x['date'], reverse=True)

    summary = {
        'total_sales_value':  total_sales,
        'total_purchase_cost': total_purchase,
        'total_gst':          total_gst,
        'total_profit':       total_profit,
        'profit_percentage':  (total_profit / total_sales * 100) if total_sales else 0,
        'stock_valuation':    0.0,
        'total_transactions': len(rows),
    }
    return rows, summary


# ─────────────────────────────────────────────────────────────────────────────
# VIEW
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def financial_report(request):
    start_date     = request.GET.get('start_date', '').strip()
    end_date       = request.GET.get('end_date', '').strip()
    product_id     = request.GET.get('product_id', '')
    product_search = request.GET.get('product_search', '')
    page_number    = request.GET.get('page', 1)

    lists = _build_queries(start_date, end_date, product_id, product_search)
    rows, summary = _build_financial_data(*lists)

    from django.core.paginator import Paginator
    page_obj = Paginator(rows, 50).get_page(page_number)

    return render(request, 'reports/financial_report.html', {
        'financial_data':   page_obj,
        'summary':          summary,
        'start_date':       start_date,
        'end_date':         end_date,
        'selected_product': product_id,
        'product_search':   product_search,
    })


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PDF
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def export_financial_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO

    start_date  = request.GET.get('start_date', '').strip()
    end_date    = request.GET.get('end_date', '').strip()
    product_id  = request.GET.get('product_id', '')
    product_search = request.GET.get('product_search', '')

    lists = _build_queries(start_date, end_date, product_id, product_search)
    rows, summary = _build_financial_data(*lists)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20,
                            topMargin=25, bottomMargin=18)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=14,
                                 textColor=colors.HexColor('#1a237e'), alignment=1)
    elements.append(Paragraph('Financial Report - Profit Analysis', title_style))
    elements.append(Spacer(1, 8))

    if start_date and end_date:
        elements.append(Paragraph(
            f"Period: {start_date} to {end_date}",
            styles['Normal']))
    elif start_date:
        elements.append(Paragraph(f"From: {start_date}", styles['Normal']))
    elif end_date:
        elements.append(Paragraph(f"Up to: {end_date}", styles['Normal']))
    elements.append(Spacer(1, 8))

    # Summary row
    elements.append(Paragraph(
        f"Total Sales: ₹{summary['total_sales_value']:.2f}  |  "
        f"Total Purchase: ₹{summary['total_purchase_cost']:.2f}  |  "
        f"Total GST: ₹{summary['total_gst']:.2f}  |  "
        f"Profit: ₹{summary['total_profit']:.2f}  |  "
        f"Transactions: {summary['total_transactions']}",
        styles['Normal']))
    elements.append(Spacer(1, 10))

    data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Batch',
             'Qty', 'P.Rate', 'S.Rate', 'GST', 'P.Cost', 'S.Value', 'Profit']]

    for r in rows:
        d = r['date']
        date_str = d.strftime('%d-%m-%Y') if hasattr(d, 'strftime') else str(d)
        data.append([
            date_str,
            r['type'],
            str(r['invoice_no'])[:12],
            str(r['customer'])[:15],
            str(r['product_name'])[:18],
            str(r['batch_no'])[:10],
            f"{r['quantity']:.0f}",
            f"{r['purchase_rate']:.2f}",
            f"{r['sale_rate']:.2f}",
            f"{r['gst_amount']:.2f}",
            f"{r['purchase_cost']:.2f}",
            f"{r['sales_value']:.2f}",
            f"{r['profit']:.2f}",
        ])

    # Totals row
    data.append(['', '', '', '', '', '', '', '', '', 'TOTAL',
                 f"{summary['total_purchase_cost']:.2f}",
                 f"{summary['total_sales_value']:.2f}",
                 f"{summary['total_profit']:.2f}"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        ('FONTSIZE',   (0, 1), (-1, -1), 7),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.pdf"')
    return response


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def export_financial_excel(request):
    start_date     = request.GET.get('start_date', '').strip()
    end_date       = request.GET.get('end_date', '').strip()
    product_id     = request.GET.get('product_id', '')
    product_search = request.GET.get('product_search', '')

    lists = _build_queries(start_date, end_date, product_id, product_search)
    rows, summary = _build_financial_data(*lists)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Financial Report'

    hdr_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    tot_fill = PatternFill(start_color='e8eaf6', end_color='e8eaf6', fill_type='solid')
    tot_font = Font(bold=True, size=11)
    border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'))

    # Period info
    ws['A1'] = f"Financial Report  |  Period: {start_date or 'All'} to {end_date or 'All'}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'] = (f"Total Sales: ₹{summary['total_sales_value']:.2f}  |  "
                f"Total Purchase: ₹{summary['total_purchase_cost']:.2f}  |  "
                f"GST: ₹{summary['total_gst']:.2f}  |  "
                f"Profit: ₹{summary['total_profit']:.2f}  |  "
                f"Transactions: {summary['total_transactions']}")
    ws['A2'].font = Font(size=10)

    headers = ['Date', 'Type', 'Invoice No', 'Party', 'Product', 'Company',
               'Batch', 'Qty', 'MRP', 'Purchase Rate', 'Sale Rate',
               'CGST%', 'SGST%', 'GST Amount', 'Purchase Cost',
               'Sales Value', 'Profit', 'Profit %']

    HDR_ROW = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center'); c.border = border

    row_num = HDR_ROW + 1
    for r in rows:
        d = r['date']
        date_str = d.strftime('%d-%m-%Y') if hasattr(d, 'strftime') else str(d)
        vals = [
            date_str, r['type'], r['invoice_no'], r['customer'],
            r['product_name'], r['company'], r['batch_no'],
            r['quantity'], r['mrp'], r['purchase_rate'], r['sale_rate'],
            r['cgst'], r['sgst'], r['gst_amount'],
            r['purchase_cost'], r['sales_value'],
            r['profit'], r['profit_percentage'],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.border = border
            if col >= 8:
                c.alignment = Alignment(horizontal='right')
        row_num += 1

    # Totals row
    tot_vals = ['TOTAL', '', '', '', '', '', '',
                '', '', '', '', '', '',
                round(summary['total_gst'], 2),
                round(summary['total_purchase_cost'], 2),
                round(summary['total_sales_value'], 2),
                round(summary['total_profit'], 2),
                round(summary['profit_percentage'], 2)]
    for col, v in enumerate(tot_vals, 1):
        c = ws.cell(row=row_num, column=col, value=v)
        c.fill = tot_fill; c.font = tot_font; c.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.xlsx"')
    wb.save(response)
    return response

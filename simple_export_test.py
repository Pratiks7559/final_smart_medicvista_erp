"""
Simple test to check Excel and PDF export functionality
Run with: python manage.py shell < simple_export_test.py
"""

print("=" * 60)
print("TESTING FINANCIAL REPORT EXPORTS")
print("=" * 60)

from core.models import SalesMaster, PurchaseMaster, CustomerChallanMaster, SupplierChallanMaster
from datetime import datetime

# Check data availability
print("\n1. Checking Data Availability...")
sales_count = SalesMaster.objects.count()
purchase_count = PurchaseMaster.objects.count()
customer_challan_count = CustomerChallanMaster.objects.count()
supplier_challan_count = SupplierChallanMaster.objects.count()

print(f"   Sales Records: {sales_count}")
print(f"   Purchase Records: {purchase_count}")
print(f"   Customer Challan Records: {customer_challan_count}")
print(f"   Supplier Challan Records: {supplier_challan_count}")

if sales_count == 0 and purchase_count == 0:
    print("\n   ❌ No data found! Cannot test exports.")
    exit()

# Test Excel Export
print("\n2. Testing Excel Export...")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"
    
    # Headers
    headers = ['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    row_num = 2
    data_rows = 0
    
    # Get first 5 sales
    sales = SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]
    print(f"   Processing {sales.count()} sales records...")
    
    for sale in sales:
        try:
            ws.cell(row=row_num, column=1).value = sale.sale_entry_date.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=2).value = 'Sale'
            ws.cell(row=row_num, column=3).value = sale.sales_invoice_no.sales_invoice_no
            ws.cell(row=row_num, column=4).value = sale.customerid.customer_name
            ws.cell(row=row_num, column=5).value = sale.product_name
            ws.cell(row=row_num, column=6).value = float(sale.sale_quantity)
            ws.cell(row=row_num, column=7).value = float(sale.sale_rate)
            ws.cell(row=row_num, column=8).value = float(sale.sale_quantity) * float(sale.sale_rate)
            
            print(f"      ✓ Row {row_num}: {sale.sales_invoice_no.sales_invoice_no} - {sale.product_name}")
            row_num += 1
            data_rows += 1
        except Exception as e:
            print(f"      ✗ Error: {e}")
    
    # Get first 5 purchases
    purchases = PurchaseMaster.objects.select_related('product_supplierid')[:5]
    print(f"   Processing {purchases.count()} purchase records...")
    
    for purchase in purchases:
        try:
            ws.cell(row=row_num, column=1).value = purchase.purchase_entry_date.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=2).value = 'Purchase'
            ws.cell(row=row_num, column=3).value = purchase.product_invoice_no
            ws.cell(row=row_num, column=4).value = purchase.product_supplierid.supplier_name
            ws.cell(row=row_num, column=5).value = purchase.product_name
            ws.cell(row=row_num, column=6).value = float(purchase.product_quantity)
            ws.cell(row=row_num, column=7).value = float(purchase.product_purchase_rate)
            ws.cell(row=row_num, column=8).value = float(purchase.product_quantity) * float(purchase.product_purchase_rate)
            
            print(f"      ✓ Row {row_num}: {purchase.product_invoice_no} - {purchase.product_name}")
            row_num += 1
            data_rows += 1
        except Exception as e:
            print(f"      ✗ Error: {e}")
    
    # Save Excel
    excel_file = "test_export.xlsx"
    wb.save(excel_file)
    
    import os
    file_size = os.path.getsize(excel_file)
    
    print(f"\n   ✅ Excel Export Success!")
    print(f"      File: {excel_file}")
    print(f"      Data Rows: {data_rows}")
    print(f"      File Size: {file_size} bytes")
    
    if data_rows == 0:
        print("      ⚠️  WARNING: No data rows added!")
    
except Exception as e:
    print(f"\n   ❌ Excel Export Failed: {e}")
    import traceback
    traceback.print_exc()

# Test PDF Export
print("\n3. Testing PDF Export...")
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Table data
    data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']]
    
    data_rows = 0
    
    # Get first 5 sales
    sales = SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]
    print(f"   Processing {sales.count()} sales records...")
    
    for sale in sales:
        try:
            data.append([
                sale.sale_entry_date.strftime('%d-%m-%Y'),
                'Sale',
                sale.sales_invoice_no.sales_invoice_no,
                sale.customerid.customer_name[:20],
                sale.product_name[:20],
                f"{float(sale.sale_quantity):.1f}",
                f"{float(sale.sale_rate):.2f}",
                f"{float(sale.sale_quantity) * float(sale.sale_rate):.2f}"
            ])
            print(f"      ✓ Added: {sale.sales_invoice_no.sales_invoice_no}")
            data_rows += 1
        except Exception as e:
            print(f"      ✗ Error: {e}")
    
    # Get first 5 purchases
    purchases = PurchaseMaster.objects.select_related('product_supplierid')[:5]
    print(f"   Processing {purchases.count()} purchase records...")
    
    for purchase in purchases:
        try:
            data.append([
                purchase.purchase_entry_date.strftime('%d-%m-%Y'),
                'Purchase',
                purchase.product_invoice_no,
                purchase.product_supplierid.supplier_name[:20],
                purchase.product_name[:20],
                f"{float(purchase.product_quantity):.1f}",
                f"{float(purchase.product_purchase_rate):.2f}",
                f"{float(purchase.product_quantity) * float(purchase.product_purchase_rate):.2f}"
            ])
            print(f"      ✓ Added: {purchase.product_invoice_no}")
            data_rows += 1
        except Exception as e:
            print(f"      ✗ Error: {e}")
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Save PDF
    buffer.seek(0)
    pdf_file = "test_export.pdf"
    with open(pdf_file, 'wb') as f:
        f.write(buffer.read())
    
    import os
    file_size = os.path.getsize(pdf_file)
    
    print(f"\n   ✅ PDF Export Success!")
    print(f"      File: {pdf_file}")
    print(f"      Data Rows: {data_rows}")
    print(f"      File Size: {file_size} bytes")
    
    if data_rows == 0:
        print("      ⚠️  WARNING: No data rows added!")
    
except Exception as e:
    print(f"\n   ❌ PDF Export Failed: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("TEST COMPLETED")
print("=" * 60)
print("\nCheck the generated files:")
print("  - test_export.xlsx")
print("  - test_export.pdf")

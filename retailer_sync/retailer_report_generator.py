"""
retailer_report_generator.py
-----------------------------
Generates PDF and Excel reports on the retailer machine from JSON data
received from the wholesaler API.

Data flow
---------
    Wholesaler MySQL
        └─► GET /api/retailer/request-data/<id>/  (JSON)
                └─► retailer_sync_service.get_request_data()
                        └─► RetailerCacheDB.save_report_data()
                                └─► ReportGenerator.generate_pdf()
                                └─► ReportGenerator.generate_excel()

No wholesaler database credentials are used here.
No PDF or Excel files are transferred over the network.
Reports are created locally and saved to the output directory.

Dependencies (already in wholesaler requirements.txt, install on retailer too):
    pip install reportlab openpyxl
"""

import os
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger('retailer_sync')


class ReportGenerator:
    """
    Generates PDF and Excel from a report data dict.

    Parameters
    ----------
    output_dir : str | Path
        Directory where generated files are saved.
        Created automatically if it does not exist.
    """

    def __init__(self, output_dir: str = 'retailer_reports'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public interface
    # ------------------------------------------------------------------

    def generate(self, report_data: dict) -> dict:
        """
        Generate PDF, Excel, and CSV for a report.

        Parameters
        ----------
        report_data : dict
            As returned by RequestSyncService.get_request_data() or
            RetailerCacheDB.get_report_data().
            Keys: request_id, request_type, from_date, to_date,
                  generated_at, data (list of row dicts)

        Returns
        -------
        {
            'ok': bool,
            'pdf_path': str | None,
            'excel_path': str | None,
            'csv_path': str | None,
            'error': str | None,
        }
        """
        try:
            pdf_path   = self.generate_pdf(report_data)
            excel_path = self.generate_excel(report_data)
            csv_path   = self.generate_csv(report_data)
            return {'ok': True, 'pdf_path': str(pdf_path),
                    'excel_path': str(excel_path), 'csv_path': str(csv_path), 'error': None}
        except Exception as e:
            logger.exception("Report generation failed for request_id=%s",
                             report_data.get('request_id'))
            return {'ok': False, 'pdf_path': None, 'excel_path': None, 'csv_path': None, 'error': str(e)}

    def generate_pdf(self, report_data: dict) -> Path:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        rtype      = report_data['request_type']
        from_date  = report_data['from_date']
        to_date    = report_data['to_date']
        rows       = report_data.get('data', [])
        filename   = self._filename(report_data, 'pdf')
        path       = self.output_dir / filename

        doc = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            leftMargin=10*mm, rightMargin=10*mm,
            topMargin=12*mm, bottomMargin=12*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                     fontSize=13, spaceAfter=4)
        sub_style   = ParagraphStyle('sub',   parent=styles['Normal'],
                                     fontSize=8, spaceAfter=8, textColor=colors.grey)
        cell_style  = ParagraphStyle('cell',  parent=styles['Normal'], fontSize=7)

        header_text = f"{rtype} Report"
        sub_text    = f"Period: {from_date} to {to_date}  |  Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

        cols, data_rows = _get_columns_and_rows(rtype, rows)

        # Build table data
        table_data = [cols]
        for r in data_rows:
            table_data.append([Paragraph(str(v), cell_style) for v in r])

        col_count  = len(cols)
        col_width  = (landscape(A4)[0] - 20*mm) / col_count

        tbl = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor('#4a4e69')),
            ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, 0),  7),
            ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE',    (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID',        (0, 0), (-1, -1), 0.3, colors.HexColor('#d1d5db')),
            ('TOPPADDING',  (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING',(0, 0),(-1, -1), 2),
        ]))

        # Summary row
        total_amount = sum(
            float(r.get('total_amount', 0) or 0) for r in rows
        )
        summary = Paragraph(
            f"Total Records: {len(rows)}  |  Total Amount: ₹{total_amount:,.2f}",
            sub_style,
        )

        story = [
            Paragraph(header_text, title_style),
            Paragraph(sub_text, sub_style),
            tbl,
            Spacer(1, 4*mm),
            summary,
        ]
        doc.build(story)
        logger.info("PDF generated: %s (%d rows)", path, len(rows))
        return path

    def generate_excel(self, report_data: dict) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )

        rtype     = report_data['request_type']
        from_date = report_data['from_date']
        to_date   = report_data['to_date']
        rows      = report_data.get('data', [])
        filename  = self._filename(report_data, 'xlsx')
        path      = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = rtype

        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{rtype} Report  |  {from_date} to {to_date}"
        ws['A1'].font      = Font(bold=True, size=12, color='4a4e69')
        ws['A1'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
        ws['A2'].font      = Font(size=9, color='6b7280')

        cols, data_rows = _get_columns_and_rows(rtype, rows)

        # Header row (row 4)
        hdr_fill = PatternFill('solid', fgColor='4a4e69')
        hdr_font = Font(bold=True, color='FFFFFF', size=9)
        thin     = Side(style='thin', color='d1d5db')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border

        # Data rows starting at row 5
        alt_fill = PatternFill('solid', fgColor='f3f4f6')
        for row_idx, row in enumerate(data_rows, 5):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = border
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.font      = Font(size=9)
                if fill:
                    cell.fill = fill

        # Auto-width
        for col_idx, col in enumerate(ws.iter_cols(min_row=4, max_row=4 + len(data_rows)), 1):
            col_letter = chr(64 + col_idx)  # A, B, C, etc.
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) 
                 for r in range(4, 4 + len(data_rows) + 1)),
                default=8
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        # Summary row
        summary_row = len(data_rows) + 6
        ws.cell(row=summary_row, column=1, value=f"Total Records: {len(rows)}")
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=9)

        total_amount = sum(float(r.get('total_amount', 0) or 0) for r in rows)
        ws.cell(row=summary_row, column=2,
                value=f"Total Amount: ₹{total_amount:,.2f}")
        ws.cell(row=summary_row, column=2).font = Font(bold=True, size=9)

        wb.save(str(path))
        logger.info("Excel generated: %s (%d rows)", path, len(rows))
        return path

    def generate_csv(self, report_data: dict) -> Path:
        """Generate CSV file with summary footer."""
        import csv
        
        rtype     = report_data['request_type']
        from_date = report_data['from_date']
        to_date   = report_data['to_date']
        rows      = report_data.get('data', [])
        filename  = self._filename(report_data, 'csv')
        path      = self.output_dir / filename

        cols, data_rows = _get_columns_and_rows(rtype, rows)

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # Header info
            writer.writerow([f"{rtype} Report - {from_date} to {to_date}"])
            writer.writerow([f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"])
            writer.writerow([])  # blank line
            
            # Column headers
            writer.writerow(cols)
            
            # Data rows
            writer.writerows(data_rows)
            
            # Summary footer
            writer.writerow([])  # blank line
            total_amount = sum(float(r.get('total_amount', 0) or 0) for r in rows)
            writer.writerow([f"Total Records: {len(rows)}", f"Total Amount: ₹{total_amount:,.2f}"])

        logger.info("CSV generated: %s (%d rows)", path, len(rows))
        return path

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _filename(self, report_data: dict, ext: str) -> str:
        ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
        rid   = report_data.get('request_id', 'X')
        rtype = report_data.get('request_type', 'RPT')
        fd    = str(report_data.get('from_date', '')).replace('-', '')
        td    = str(report_data.get('to_date',   '')).replace('-', '')
        return f"{rtype}_{fd}_{td}_REQ{rid}_{ts}.{ext}"


# ---------------------------------------------------------------------------
# Column definitions per report type
# ---------------------------------------------------------------------------

_SALES_COLS = [
    'Invoice No', 'Date', 'Customer', 'Type',
    'Product', 'Company', 'Packing', 'Batch', 'Expiry',
    'MRP', 'Rate', 'Qty', 'Free', 'Discount',
    'CGST', 'SGST', 'Total',
]
_SALES_KEYS = [
    'invoice_no', 'invoice_date', 'customer_name', 'customer_type',
    'product_name', 'product_company', 'product_packing', 'batch_no', 'expiry',
    'mrp', 'sale_rate', 'quantity', 'free_qty', 'discount',
    'cgst', 'sgst', 'total_amount',
]

_PURCHASE_COLS = [
    'Invoice No', 'Date', 'Supplier',
    'Product', 'Company', 'Packing', 'Batch', 'Expiry',
    'MRP', 'Rate', 'Qty', 'Free', 'Discount',
    'CGST', 'SGST', 'Total',
]
_PURCHASE_KEYS = [
    'invoice_no', 'invoice_date', 'supplier_name',
    'product_name', 'product_company', 'product_packing', 'batch_no', 'expiry',
    'mrp', 'purchase_rate', 'quantity', 'free_qty', 'discount',
    'cgst', 'sgst', 'total_amount',
]

_STOCK_COLS = [
    'Product', 'Company', 'Packing', 'HSN',
    'Batch', 'Expiry', 'MRP', 'Purchase Rate',
    'Stock', 'Free Stock', 'Total Stock',
    'Rate A', 'Rate B', 'Rate C', 'Status',
]
_STOCK_KEYS = [
    'product_name', 'product_company', 'product_packing', 'hsn',
    'batch_no', 'expiry', 'mrp', 'purchase_rate',
    'current_stock', 'free_stock', 'total_stock',
    'rate_a', 'rate_b', 'rate_c', 'expiry_status',
]


def _get_keys_for_type(rtype: str) -> list:
    if rtype == 'SALES':
        return _SALES_KEYS
    if rtype == 'PURCHASE':
        return _PURCHASE_KEYS
    return _STOCK_KEYS


def _get_columns_and_rows(rtype: str, rows: list):
    if rtype == 'SALES':
        cols = _SALES_COLS
    elif rtype == 'PURCHASE':
        cols = _PURCHASE_COLS
    else:  # STOCK
        cols = _STOCK_COLS

    keys = _get_keys_for_type(rtype)
    data_rows = [[r.get(k, '') for k in keys] for r in rows]
    return cols, data_rows

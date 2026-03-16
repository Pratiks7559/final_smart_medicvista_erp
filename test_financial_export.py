"""
Test script to check if financial report exports are working
"""
import os
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'WebsiteHostingService.settings')
django.setup()

from core.models import SalesMaster, PurchaseMaster, CustomerChallanMaster, SupplierChallanMaster
from datetime import datetime, timedelta

def test_data_availability():
    """Check if data exists in database"""
    print("=" * 60)
    print("TESTING DATA AVAILABILITY")
    print("=" * 60)
    
    sales_count = SalesMaster.objects.count()
    purchase_count = PurchaseMaster.objects.count()
    customer_challan_count = CustomerChallanMaster.objects.count()
    supplier_challan_count = SupplierChallanMaster.objects.count()
    
    print(f"Sales Records: {sales_count}")
    print(f"Purchase Records: {purchase_count}")
    print(f"Customer Challan Records: {customer_challan_count}")
    print(f"Supplier Challan Records: {supplier_challan_count}")
    print()
    
    if sales_count == 0 and purchase_count == 0:
        print("⚠️  WARNING: No data found in database!")
        return False
    
    return True

def test_excel_export():
    """Test Excel export functionality"""
    print("=" * 60)
    print("TESTING EXCEL EXPORT")
    print("=" * 60)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get sample data
        sales_query = SalesMaster.objects.select_related('productid', 'sales_invoice_no', 'customerid')[:10]
        purchase_query = PurchaseMaster.objects.select_related('productid', 'product_invoiceid', 'product_supplierid')[:10]
        
        print(f"Fetched {sales_query.count()} sales records")
        print(f"Fetched {purchase_query.count()} purchase records")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Financial Report"
        
        # Headers
        headers = ['Date', 'Type', 'Invoice No', 'Party', 'Product', 'Company', 'Batch', 'Qty', 
                   'MRP', 'Purchase Rate', 'Sale Rate', 'CGST', 'SGST', 'GST Amount', 
                   'Purchase Cost', 'Sales Value', 'Profit', 'Profit %']
        
        # Header styling
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row_num = 2
        total_rows_added = 0
        
        # Process Sales
        print("\nProcessing Sales...")
        for sale in sales_query:
            try:
                purchase = PurchaseMaster.objects.filter(
                    productid=sale.productid, 
                    product_batch_no=sale.product_batch_no
                ).first()
                
                purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
                quantity = float(sale.sale_quantity)
                sale_rate = float(sale.sale_rate)
                purchase_cost = purchase_rate * quantity
                sales_value = sale_rate * quantity
                gst_amount = sales_value * (float(sale.sale_cgst) + float(sale.sale_sgst)) / 100
                profit = sales_value - purchase_cost
                profit_pct = (profit / sales_value * 100) if sales_value > 0 else 0
                
                row_data = [
                    sale.sale_entry_date.strftime('%d-%m-%Y'), 
                    'Sale', 
                    sale.sales_invoice_no.sales_invoice_no,
                    sale.customerid.customer_name, 
                    sale.product_name, 
                    sale.product_company, 
                    sale.product_batch_no,
                    quantity, 
                    float(sale.product_MRP), 
                    purchase_rate, 
                    sale_rate, 
                    float(sale.sale_cgst), 
                    float(sale.sale_sgst),
                    gst_amount, 
                    purchase_cost, 
                    sales_value, 
                    profit, 
                    profit_pct
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num >= 8:
                        cell.alignment = Alignment(horizontal='right')
                
                print(f"  ✓ Added Sale: {sale.sales_invoice_no.sales_invoice_no} - {sale.product_name}")
                row_num += 1
                total_rows_added += 1
                
            except Exception as e:
                print(f"  ✗ Error processing sale: {e}")
                continue
        
        # Process Purchases
        print("\nProcessing Purchases...")
        for purchase in purchase_query:
            try:
                quantity = float(purchase.product_quantity)
                purchase_rate = float(purchase.product_purchase_rate)
                purchase_cost = purchase_rate * quantity
                gst_amount = purchase_cost * (float(purchase.CGST) + float(purchase.SGST)) / 100
                profit = -purchase_cost
                
                row_data = [
                    purchase.purchase_entry_date.strftime('%d-%m-%Y'), 
                    'Purchase', 
                    purchase.product_invoice_no,
                    purchase.product_supplierid.supplier_name, 
                    purchase.product_name, 
                    purchase.product_company, 
                    purchase.product_batch_no,
                    quantity, 
                    float(purchase.product_MRP), 
                    purchase_rate, 
                    0, 
                    float(purchase.CGST), 
                    float(purchase.SGST),
                    gst_amount, 
                    purchase_cost, 
                    0, 
                    profit, 
                    0
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num >= 8:
                        cell.alignment = Alignment(horizontal='right')
                
                print(f"  ✓ Added Purchase: {purchase.product_invoice_no} - {purchase.product_name}")
                row_num += 1
                total_rows_added += 1
                
            except Exception as e:
                print(f"  ✗ Error processing purchase: {e}")
                continue
        
        # Adjust column widths
        for col in range(1, 19):
            ws.column_dimensions[get_column_letter(col)].width = 14
        
        # Save file
        test_file = "test_financial_report.xlsx"
        wb.save(test_file)
        
        print(f"\n✅ Excel file created successfully: {test_file}")
        print(f"✅ Total data rows added: {total_rows_added}")
        print(f"✅ File size: {os.path.getsize(test_file)} bytes")
        
        if total_rows_added == 0:
            print("⚠️  WARNING: No data rows were added to Excel!")
            return False
        
        return True
        
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_pdf_export():
    """Test PDF export functionality"""
    print("\n" + "=" * 60)
    print("TESTING PDF EXPORT")
    print("=" * 60)
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO
        
        # Get sample data
        sales_query = SalesMaster.objects.select_related('productid', 'sales_invoice_no', 'customerid')[:10]
        purchase_query = PurchaseMaster.objects.select_related('productid', 'product_invoiceid', 'product_supplierid')[:10]
        
        print(f"Fetched {sales_query.count()} sales records")
        print(f"Fetched {purchase_query.count()} purchase records")
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1a237e'), alignment=1)
        elements.append(Paragraph('Test Financial Report', title_style))
        elements.append(Spacer(1, 12))
        
        # Table data
        data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Batch', 'Qty', 'P.Rate', 'S.Rate', 'GST', 'Profit']]
        
        all_txns = []
        
        # Collect Sales
        print("\nProcessing Sales for PDF...")
        for sale in sales_query:
            try:
                purchase = PurchaseMaster.objects.filter(
                    productid=sale.productid, 
                    product_batch_no=sale.product_batch_no
                ).first()
                
                purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
                quantity = float(sale.sale_quantity)
                sale_rate = float(sale.sale_rate)
                purchase_cost = purchase_rate * quantity
                sales_value = sale_rate * quantity
                gst_amount = sales_value * (float(sale.sale_cgst) + float(sale.sale_sgst)) / 100
                profit = sales_value - purchase_cost
                
                all_txns.append({
                    'date': sale.sale_entry_date, 
                    'type': 'Sale', 
                    'invoice': sale.sales_invoice_no.sales_invoice_no,
                    'party': sale.customerid.customer_name, 
                    'product': sale.product_name[:15], 
                    'batch': sale.product_batch_no,
                    'qty': quantity, 
                    'p_rate': purchase_rate, 
                    's_rate': sale_rate, 
                    'gst': gst_amount, 
                    'profit': profit, 
                    'sales': sales_value
                })
                
                print(f"  ✓ Added Sale: {sale.sales_invoice_no.sales_invoice_no}")
                
            except Exception as e:
                print(f"  ✗ Error processing sale: {e}")
                continue
        
        # Collect Purchases
        print("\nProcessing Purchases for PDF...")
        for purchase in purchase_query:
            try:
                quantity = float(purchase.product_quantity)
                purchase_rate = float(purchase.product_purchase_rate)
                purchase_cost = purchase_rate * quantity
                gst_amount = purchase_cost * (float(purchase.CGST) + float(purchase.SGST)) / 100
                profit = -purchase_cost
                
                all_txns.append({
                    'date': purchase.purchase_entry_date, 
                    'type': 'Purchase', 
                    'invoice': purchase.product_invoice_no,
                    'party': purchase.product_supplierid.supplier_name, 
                    'product': purchase.product_name[:15], 
                    'batch': purchase.product_batch_no,
                    'qty': quantity, 
                    'p_rate': purchase_rate, 
                    's_rate': 0, 
                    'gst': gst_amount, 
                    'profit': profit, 
                    'sales': 0
                })
                
                print(f"  ✓ Added Purchase: {purchase.product_invoice_no}")
                
            except Exception as e:
                print(f"  ✗ Error processing purchase: {e}")
                continue
        
        # Sort and add to table
        total_profit = 0.0
        if all_txns:
            all_txns.sort(key=lambda x: x['date'], reverse=True)
            for txn in all_txns:
                data.append([
                    txn['date'].strftime('%d-%m-%Y'),
                    txn['type'],
                    str(txn['invoice'])[:10],
                    str(txn['party'])[:15],
                    str(txn['product'])[:12],
                    str(txn['batch'])[:8],
                    f"{txn['qty']:.1f}",
                    f"{txn['p_rate']:.2f}",
                    f"{txn['s_rate']:.2f}",
                    f"{txn['gst']:.2f}",
                    f"{txn['profit']:.2f}"
                ])
                total_profit += txn['profit']
        
        # Add summary row
        data.append(['', '', '', '', '', '', '', '', '', 'TOTAL:', f'{total_profit:.2f}'])
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        # Save to file
        buffer.seek(0)
        test_file = "test_financial_report.pdf"
        with open(test_file, 'wb') as f:
            f.write(buffer.read())
        
        print(f"\n✅ PDF file created successfully: {test_file}")
        print(f"✅ Total data rows added: {len(all_txns)}")
        print(f"✅ File size: {os.path.getsize(test_file)} bytes")
        
        if len(all_txns) == 0:
            print("⚠️  WARNING: No data rows were added to PDF!")
            return False
        
        return True
        
    except Exception as e:
        print(f"❌ PDF export failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("\n" + "🔍 FINANCIAL REPORT EXPORT TEST" + "\n")
    
    # Test 1: Check data availability
    data_exists = test_data_availability()
    
    if not data_exists:
        print("\n❌ Cannot proceed with export tests - No data in database")
        exit(1)
    
    # Test 2: Excel Export
    excel_success = test_excel_export()
    
    # Test 3: PDF Export
    pdf_success = test_pdf_export()
    
    # Final Summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"Data Availability: {'✅ PASS' if data_exists else '❌ FAIL'}")
    print(f"Excel Export: {'✅ PASS' if excel_success else '❌ FAIL'}")
    print(f"PDF Export: {'✅ PASS' if pdf_success else '❌ FAIL'}")
    print("=" * 60)
    
    if excel_success and pdf_success:
        print("\n🎉 All tests passed! Check the generated files:")
        print("   - test_financial_report.xlsx")
        print("   - test_financial_report.pdf")
    else:
        print("\n⚠️  Some tests failed. Check the error messages above.")

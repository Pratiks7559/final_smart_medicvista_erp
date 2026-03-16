import os
import sys
import django

# Add the project directory to the Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pharmamgmt.settings')
django.setup()

from core.models import SalesMaster, PurchaseMaster
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 70)
print("TESTING FINANCIAL EXPORT - DATA CHECK")
print("=" * 70)

# Check database
print("\n1. Checking Database...")
sales_count = SalesMaster.objects.count()
purchase_count = PurchaseMaster.objects.count()
print(f"   Sales Records: {sales_count}")
print(f"   Purchase Records: {purchase_count}")

if sales_count == 0 and purchase_count == 0:
    print("\n❌ ERROR: No data in database!")
    sys.exit(1)

# Test Excel Export
print("\n2. Testing Excel Export...")
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"
    
    # Headers
    headers = ['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    row_num = 2
    data_rows = 0
    
    # Get sales data
    sales = SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]
    print(f"   Fetching {sales.count()} sales records...")
    
    for sale in sales:
        try:
            ws.cell(row=row_num, column=1, value=sale.sale_entry_date.strftime('%d-%m-%Y'))
            ws.cell(row=row_num, column=2, value='Sale')
            ws.cell(row=row_num, column=3, value=sale.sales_invoice_no.sales_invoice_no)
            ws.cell(row=row_num, column=4, value=sale.customerid.customer_name)
            ws.cell(row=row_num, column=5, value=sale.product_name)
            ws.cell(row=row_num, column=6, value=float(sale.sale_quantity))
            ws.cell(row=row_num, column=7, value=float(sale.sale_rate))
            ws.cell(row=row_num, column=8, value=float(sale.sale_quantity) * float(sale.sale_rate))
            
            print(f"   ✓ Row {row_num}: {sale.product_name} - Qty: {sale.sale_quantity}")
            row_num += 1
            data_rows += 1
        except Exception as e:
            print(f"   ✗ Error in row {row_num}: {e}")
    
    # Get purchase data
    purchases = PurchaseMaster.objects.select_related('product_supplierid')[:5]
    print(f"   Fetching {purchases.count()} purchase records...")
    
    for purchase in purchases:
        try:
            ws.cell(row=row_num, column=1, value=purchase.purchase_entry_date.strftime('%d-%m-%Y'))
            ws.cell(row=row_num, column=2, value='Purchase')
            ws.cell(row=row_num, column=3, value=purchase.product_invoice_no)
            ws.cell(row=row_num, column=4, value=purchase.product_supplierid.supplier_name)
            ws.cell(row=row_num, column=5, value=purchase.product_name)
            ws.cell(row=row_num, column=6, value=float(purchase.product_quantity))
            ws.cell(row=row_num, column=7, value=float(purchase.product_purchase_rate))
            ws.cell(row=row_num, column=8, value=float(purchase.product_quantity) * float(purchase.product_purchase_rate))
            
            print(f"   ✓ Row {row_num}: {purchase.product_name} - Qty: {purchase.product_quantity}")
            row_num += 1
            data_rows += 1
        except Exception as e:
            print(f"   ✗ Error in row {row_num}: {e}")
    
    # Save file
    filename = "test_export.xlsx"
    wb.save(filename)
    file_size = os.path.getsize(filename)
    
    print(f"\n   ✅ Excel file created: {filename}")
    print(f"   ✅ Data rows added: {data_rows}")
    print(f"   ✅ File size: {file_size} bytes")
    
    if data_rows == 0:
        print("   ⚠️  WARNING: No data rows in Excel!")
    
except Exception as e:
    print(f"\n   ❌ Excel export failed: {e}")
    import traceback
    traceback.print_exc()

# Test PDF Export
print("\n3. Testing PDF Export...")
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    
    # Table data
    data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']]
    
    pdf_rows = 0
    
    # Get sales
    sales = SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]
    print(f"   Fetching {sales.count()} sales records...")
    
    for sale in sales:
        try:
            data.append([
                sale.sale_entry_date.strftime('%d-%m-%Y'),
                'Sale',
                sale.sales_invoice_no.sales_invoice_no,
                sale.customerid.customer_name[:20],
                sale.product_name[:20],
                f"{float(sale.sale_quantity):.1f}",
                f"{float(sale.sale_rate):.2f}",
                f"{float(sale.sale_quantity) * float(sale.sale_rate):.2f}"
            ])
            print(f"   ✓ Added: {sale.product_name}")
            pdf_rows += 1
        except Exception as e:
            print(f"   ✗ Error: {e}")
    
    # Get purchases
    purchases = PurchaseMaster.objects.select_related('product_supplierid')[:5]
    print(f"   Fetching {purchases.count()} purchase records...")
    
    for purchase in purchases:
        try:
            data.append([
                purchase.purchase_entry_date.strftime('%d-%m-%Y'),
                'Purchase',
                purchase.product_invoice_no,
                purchase.product_supplierid.supplier_name[:20],
                purchase.product_name[:20],
                f"{float(purchase.product_quantity):.1f}",
                f"{float(purchase.product_purchase_rate):.2f}",
                f"{float(purchase.product_quantity) * float(purchase.product_purchase_rate):.2f}"
            ])
            print(f"   ✓ Added: {purchase.product_name}")
            pdf_rows += 1
        except Exception as e:
            print(f"   ✗ Error: {e}")
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    doc.build([table])
    
    # Save file
    buffer.seek(0)
    filename = "test_export.pdf"
    with open(filename, 'wb') as f:
        f.write(buffer.read())
    
    file_size = os.path.getsize(filename)
    
    print(f"\n   ✅ PDF file created: {filename}")
    print(f"   ✅ Data rows added: {pdf_rows}")
    print(f"   ✅ File size: {file_size} bytes")
    
    if pdf_rows == 0:
        print("   ⚠️  WARNING: No data rows in PDF!")
    
except Exception as e:
    print(f"\n   ❌ PDF export failed: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 70)
print("TEST COMPLETED - Check test_export.xlsx and test_export.pdf files")
print("=" * 70)

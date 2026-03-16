import os
import sys
import django

# Setup Django
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pharmamgmt.settings')
django.setup()

from core.models import SalesMaster, PurchaseMaster, CustomerChallanMaster, SupplierChallanMaster

print("=" * 60)
print("TESTING FINANCIAL REPORT EXPORTS")
print("=" * 60)

# Check data
print("\n1. Checking Data...")
sales_count = SalesMaster.objects.count()
purchase_count = PurchaseMaster.objects.count()
print(f"   Sales: {sales_count}")
print(f"   Purchases: {purchase_count}")

if sales_count == 0 and purchase_count == 0:
    print("\n   ERROR: No data found!")
    sys.exit(1)

# Test Excel
print("\n2. Testing Excel Export...")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    
    wb = Workbook()
    ws = wb.active
    
    # Headers
    headers = ['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col).value = h
        ws.cell(1, col).font = Font(bold=True)
    
    row = 2
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add Sales
    for sale in SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]:
        try:
            ws.cell(row, 1).value = sale.sale_entry_date.strftime('%d-%m-%Y')
            ws.cell(row, 2).value = 'Sale'
            ws.cell(row, 3).value = sale.sales_invoice_no.sales_invoice_no
            ws.cell(row, 4).value = sale.customerid.customer_name
            ws.cell(row, 5).value = sale.product_name
            ws.cell(row, 6).value = float(sale.sale_quantity)
            ws.cell(row, 7).value = float(sale.sale_rate)
            ws.cell(row, 8).value = float(sale.sale_quantity) * float(sale.sale_rate)
            
            for col in range(1, 9):
                ws.cell(row, col).border = border
            
            print(f"   Added Sale Row {row}: {sale.sales_invoice_no.sales_invoice_no}")
            row += 1
        except Exception as e:
            print(f"   Error in sale: {e}")
    
    # Add Purchases
    for purchase in PurchaseMaster.objects.select_related('product_supplierid')[:5]:
        try:
            ws.cell(row, 1).value = purchase.purchase_entry_date.strftime('%d-%m-%Y')
            ws.cell(row, 2).value = 'Purchase'
            ws.cell(row, 3).value = purchase.product_invoice_no
            ws.cell(row, 4).value = purchase.product_supplierid.supplier_name
            ws.cell(row, 5).value = purchase.product_name
            ws.cell(row, 6).value = float(purchase.product_quantity)
            ws.cell(row, 7).value = float(purchase.product_purchase_rate)
            ws.cell(row, 8).value = float(purchase.product_quantity) * float(purchase.product_purchase_rate)
            
            for col in range(1, 9):
                ws.cell(row, col).border = border
            
            print(f"   Added Purchase Row {row}: {purchase.product_invoice_no}")
            row += 1
        except Exception as e:
            print(f"   Error in purchase: {e}")
    
    # Save
    excel_file = 'test_export.xlsx'
    wb.save(excel_file)
    file_size = os.path.getsize(excel_file)
    
    print(f"\n   SUCCESS: Excel file created!")
    print(f"   File: {excel_file}")
    print(f"   Size: {file_size} bytes")
    print(f"   Data Rows: {row - 2}")
    
except Exception as e:
    print(f"\n   FAILED: {e}")
    import traceback
    traceback.print_exc()

# Test PDF
print("\n3. Testing PDF Export...")
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    
    # Table data
    data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Qty', 'Rate', 'Amount']]
    
    # Add Sales
    for sale in SalesMaster.objects.select_related('sales_invoice_no', 'customerid')[:5]:
        try:
            data.append([
                sale.sale_entry_date.strftime('%d-%m-%Y'),
                'Sale',
                sale.sales_invoice_no.sales_invoice_no,
                sale.customerid.customer_name[:20],
                sale.product_name[:20],
                f"{float(sale.sale_quantity):.1f}",
                f"{float(sale.sale_rate):.2f}",
                f"{float(sale.sale_quantity) * float(sale.sale_rate):.2f}"
            ])
            print(f"   Added Sale: {sale.sales_invoice_no.sales_invoice_no}")
        except Exception as e:
            print(f"   Error in sale: {e}")
    
    # Add Purchases
    for purchase in PurchaseMaster.objects.select_related('product_supplierid')[:5]:
        try:
            data.append([
                purchase.purchase_entry_date.strftime('%d-%m-%Y'),
                'Purchase',
                purchase.product_invoice_no,
                purchase.product_supplierid.supplier_name[:20],
                purchase.product_name[:20],
                f"{float(purchase.product_quantity):.1f}",
                f"{float(purchase.product_purchase_rate):.2f}",
                f"{float(purchase.product_quantity) * float(purchase.product_purchase_rate):.2f}"
            ])
            print(f"   Added Purchase: {purchase.product_invoice_no}")
        except Exception as e:
            print(f"   Error in purchase: {e}")
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    doc.build([table])
    
    # Save
    pdf_file = 'test_export.pdf'
    with open(pdf_file, 'wb') as f:
        f.write(buffer.getvalue())
    
    file_size = os.path.getsize(pdf_file)
    
    print(f"\n   SUCCESS: PDF file created!")
    print(f"   File: {pdf_file}")
    print(f"   Size: {file_size} bytes")
    print(f"   Data Rows: {len(data) - 1}")
    
except Exception as e:
    print(f"\n   FAILED: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("TEST COMPLETED!")
print("=" * 60)
print("\nGenerated files:")
print("  - test_export.xlsx")
print("  - test_export.pdf")
print("\nOpen these files to verify data is visible.")
